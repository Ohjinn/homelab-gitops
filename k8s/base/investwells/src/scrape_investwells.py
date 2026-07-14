#!/usr/bin/env python3
"""
investwells.com 파트너사 정보 스크래퍼 (Excel 출력)

/partners/all 페이지의 모든 회사 정보를 추출해서 엑셀(.xlsx)로 저장한다.

동작 방식 (사이트 내부 AJAX 엔드포인트 그대로 사용):
  1) POST /partners/getPtnrdetail  (idx=0 = '전체' 카테고리) → 전체 파트너 목록
  2) POST /partners/getPtnrInfo    (idx=<파트너 idx>)        → 카테고리명 + 관련뉴스

추출 항목 (회사당 6가지):
  - 회사이름      part_nm_ko
  - 카테고리      get_cate_nm  (getPtnrInfo)
  - 투자년도      inv_year
  - 홈페이지 URL  url
  - 회사소개      corp_nm_ko
  - 관련뉴스      news[] (title / url / date)  (getPtnrInfo)

결과물: investwells_partners.xlsx  (+ 원본 백업 investwells_partners.json)

의존성:
  - 네트워크: 파이썬 표준 라이브러리(urllib)만 사용 → 별도 설치 불필요
  - 엑셀:     openpyxl (없으면 자동 설치 시도)

실행:  python3 scrape_investwells.py
"""

import json
import sys
import time
import urllib.parse
import urllib.request

# Windows(특히 한글 cp949 콘솔)에서 한글·특수문자 출력 시 UnicodeEncodeError 방지.
# 콘솔/리다이렉트 스트림을 utf-8로 재설정한다. (Python 3.7+)
for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding="utf-8")
    except (AttributeError, ValueError):
        pass


# ---------------------------------------------------------------------------
# 의존성 자동 설치 (다른 PC에서도 그대로 돌아가도록)
# ---------------------------------------------------------------------------
def ensure_package(module_name, pip_name=None):
    """모듈이 없으면 pip로 자동 설치한다. 환경별 pip 제약을 순차 시도."""
    try:
        return __import__(module_name)
    except ImportError:
        pass

    import subprocess

    pip_name = pip_name or module_name
    attempts = [
        [sys.executable, "-m", "pip", "install", pip_name],
        [sys.executable, "-m", "pip", "install", "--user", pip_name],
        [sys.executable, "-m", "pip", "install", "--break-system-packages", pip_name],
    ]
    for cmd in attempts:
        try:
            print(f"[설치] {' '.join(cmd)}", file=sys.stderr)
            subprocess.check_call(cmd, stdout=sys.stderr, stderr=sys.stderr)
            return __import__(module_name)
        except (subprocess.CalledProcessError, ImportError):
            continue
    raise SystemExit(
        f"'{pip_name}' 설치에 실패했습니다. 수동으로 설치 후 다시 실행하세요:\n"
        f"    pip install {pip_name}"
    )


# ---------------------------------------------------------------------------
# 스크래핑
# ---------------------------------------------------------------------------
BASE = "https://investwells.com"
LIST_URL = f"{BASE}/partners/getPtnrdetail"
INFO_URL = f"{BASE}/partners/getPtnrInfo"

HEADERS = {
    "X-Requested-With": "XMLHttpRequest",
    "User-Agent": "Mozilla/5.0 (compatible; partner-scraper/1.0)",
    "Referer": f"{BASE}/partners/all",
}


def post_json(url, fields):
    """표준 라이브러리로 POST 요청 후 JSON 파싱."""
    data = urllib.parse.urlencode(fields).encode("utf-8")
    req = urllib.request.Request(url, data=data, headers=HEADERS, method="POST")
    with urllib.request.urlopen(req, timeout=30) as resp:
        return json.loads(resp.read().decode("utf-8"))


def get_partner_list():
    """전체 파트너 목록 (idx=0 → 전체 카테고리)."""
    return post_json(LIST_URL, {"idx": 0}).get("list", [])


def get_partner_info(idx):
    """단일 파트너의 카테고리명 + 관련뉴스."""
    return post_json(INFO_URL, {"idx": idx, "cate_idx": ""})


def scrape(delay=0.15):
    partners = get_partner_list()
    print(f"파트너 {len(partners)}개 발견", file=sys.stderr)

    results = []
    for i, p in enumerate(partners, 1):
        idx = p.get("idx")
        info = get_partner_info(idx)
        data = info.get("data") or {}
        news = [
            {
                "title": n.get("title"),
                "url": n.get("url"),
                "date": n.get("submit_date2"),
            }
            for n in (info.get("news") or [])
        ]
        results.append(
            {
                "idx": idx,
                "회사이름": p.get("part_nm_ko"),
                "카테고리": data.get("get_cate_nm") or "",
                "투자년도": p.get("inv_year"),
                "홈페이지": p.get("url"),
                "회사소개": p.get("corp_nm_ko"),
                "관련뉴스": news,
            }
        )
        print(f"[{i}/{len(partners)}] {p.get('part_nm_ko')}", file=sys.stderr)
        if delay:
            time.sleep(delay)
    return results


# ---------------------------------------------------------------------------
# 저장
# ---------------------------------------------------------------------------
def save_json(results, path="investwells_partners.json"):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    print(f"저장: {path}", file=sys.stderr)


def save_xlsx(results, path="investwells_partners.xlsx"):
    openpyxl = ensure_package("openpyxl")
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "파트너사"

    headers = ["회사이름", "카테고리", "투자년도", "홈페이지", "회사소개", "관련뉴스"]
    ws.append(headers)

    # 헤더 스타일
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in results:
        news_cell = "\n".join(
            f"{n['title']} ({n['date']}) {n['url']}" for n in r["관련뉴스"]
        )
        ws.append(
            [
                r["회사이름"],
                r["카테고리"],
                r["투자년도"],
                r["홈페이지"],
                r["회사소개"],
                news_cell,
            ]
        )

    # 열 너비 / 정렬
    widths = [16, 18, 10, 34, 55, 70]
    for col_idx, width in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    wrap = Alignment(vertical="top", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap

    ws.freeze_panes = "A2"  # 헤더 고정
    wb.save(path)
    print(f"저장: {path}", file=sys.stderr)


if __name__ == "__main__":
    data = scrape()
    save_json(data)
    save_xlsx(data)
    print(f"완료: 총 {len(data)}개 회사 → investwells_partners.xlsx", file=sys.stderr)
